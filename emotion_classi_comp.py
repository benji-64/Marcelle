import torch
import pandas as pd
import os
import time
import math
import openpyxl
import subprocess
import datetime
from transformers import AutoTokenizer, AutoModelForSequenceClassification, pipeline
 
#import gc


print("begin")


#############Préparation des données
print("donnees")
# ditp = pd.read_csv('2025_10_21_DITP.csv', encoding="utf8", sep=";")
ditp = pd.read_csv('./2025_10_21_DITP.csv', encoding="utf8", sep=";")
df=ditp[["ID expérience","Description"]]

mask = df["Description"].str.contains("amande", case=False, na=False)
df.loc[mask, "Description"] = df.loc[mask, "Description"].str.replace("amande", "amende", case=False, regex=True)


#############modeles parametres
print('parametrages')

if torch.cuda.is_available():
    device = torch.device('cuda')
elif torch.backends.mps.is_available():
    device = torch.device("mps")
else:
    device = torch.device('cpu')
    print ("GPU not found.")



###### push to github fonction
print("push to github")
def push_to_github(commit_message=None):
    """
    Pushes results to GitHub
    """
    try: 

        if commit_message is None:
            commit_message = f"Auto-update: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        subprocess.run(['git', 'add', '.'], check = True)

        subprocess.run(['git', 'commit', '-m', commit_message], check=True)

        subprocess.run(['git', 'push', 'origin', 'master'], check=True)

        print(f"Successfully pushed to GitHub: {commit_message}")

    except subprocess.CalledProcessError as e:
        print(f"Error  pushing to GitHub: {e}")




# classes = [
#     "Acceptation", "Affection", "Amour", "Amusement", "Angoisse", "Anxiété", "Appréhension", "Appartenance",
#     "Bonheur", "Calme", "Colère", "Confusion", "Déception", "Dégoût", "Dépit", "Désespoir", "Désir", "Détresse",
#     "Dignité", "Doute", "Enthousiasme", "Espérance", "Espoir", "Étonnement", "Excitation", "Expression",
#     "Fierté", "Frustration", "Gratitude", "Honte", "Humiliation", "Indignation", "Inquiétude", "Insatisfaction",
#     "Intensité", "Joie", "Maîtrise", "Mépris","Optimisme",  "Peur", "Pessimisme", "Plaisir",
#     "Regret", "Réjouissance", "Renoncement", "Respect", "Revendication","Satisfaction", "Soulagement",
#     "Surprise", "Tristesse",
#     "Résignation", "Décence", "Distinction", "Considération", "Justice procédurale", "Justice interactionnelle", "Justice restauratrice"
#     ]

#classes = ["Résignation", "Décence", "Distinction", "Justice procédurale", "Justice interactionnelle", "Justice restauratrice"
 #   ]

classes = ["Détermination", "Persevérance", "Engagement", "Résignation", "Abandon", "Égalité", "Équité", "Motivation", "Détachement", "Indifférence",
           "Apathie", "Lassitude", "Apaisement", "Tranquillité", "Résignation", "Passivité", "Découragement", "Sentiment d'injustice", "Incompréhension",
           "Arbitraire", "Considération", "Impuissance", "Incertitude", "Impatience", "Confiance", "Rassurance", "Sécurité", "Fatigue administrative",
           "Sentiment d'abandon" ]

print(classes)

batch_size =2
batches = {} #Pour nommer les dataframe avant sauvegarde

output_folder = os.path.join(os.getcwd(), "Resultats/V3")
os.makedirs(output_folder, exist_ok=True)


########################## premier modele #######################

print('modele deberta')
m='deberta'


zeroshot_classifier = pipeline("zero-shot-classification", model="MoritzLaurer/deberta-v3-large-zeroshot-v2.0",
                               device=device, batch_size=batch_size, multilabel=True)
hypothesis_template = "The emotions of this text are {}"



# Enregistrer le temps de début
start_time = time.time()
print(start_time)



save_batch = 1000
# total_rows = 100
total_rows = len(df)
num_batches = math.ceil(total_rows / save_batch)
excel_file = os.path.join(output_folder, f"{m}.xlsx")
print("11")

for batch_idx in range(num_batches):
    print(f"Processing batch {batch_idx + 1} of {num_batches}")
    start = batch_idx * save_batch
    end = min(start + save_batch, total_rows)

    batch_rows = []
    for i, row in df.iloc[start:end].iterrows():
        output = zeroshot_classifier(
            row["Description"],
            classes,
            hypothesis_template=hypothesis_template,
            multi_label=True
        )
        scores_dict = dict(zip(output["labels"], output["scores"]))
        row_result = {label: scores_dict.get(label, 0.0) for label in classes}
        row_result["ID expérience"] = row["ID expérience"]
        row_result["Description"] = row["Description"]
        batch_rows.append(row_result)

    # Convert all results to DataFrame
    df_batch = pd.DataFrame(batch_rows)[["ID expérience", "Description"] + classes]


    # Save to Excel (append or create)
    if batch_idx == 0 or not os.path.exists(excel_file):
        # First batch: write with header
        df_batch.to_excel(excel_file, index=False,  engine="openpyxl")
    else:
        # Append without writing header
        with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            # Get existing rows in sheet
            book = writer.book
            sheet = book.active
            is_empty = (sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None)
            startrow = 0 if is_empty else sheet.max_row

            # Write below existing rows
            df_batch.to_excel(writer, index=False, header=False, startrow=startrow)

            # if device.type == "mps":
            #     torch.mps.empty_cache()
            # gc.collect()

    print(f"Saved batch {batch_idx + 1} to {excel_file}")



# Enregistrer le temps de fin
end_time = time.time()

# Calculer le temps d'exécution
execution_time = end_time - start_time
print(f"Temps d'exécution: {execution_time} secondes")

times = pd.DataFrame([{
    "Modèle": m,
    "Taille de lot": batch_size,
    "Execution (s)": round(execution_time, 2),
}])

push_to_github("Deberta")


####################### Deuxieme modele ############################

print('modele roberta')
m='roberta'
excel_file = os.path.join(output_folder, f"{m}.xlsx")

zeroshot_classifier = pipeline("zero-shot-classification", model="joeddav/xlm-roberta-large-xnli",
                               device=device, batch_size=batch_size, multilabel=True)
hypothesis_template = "The emotions of this text are {}"


# Enregistrer le temps de début
start_time = time.time()
print(start_time)


for batch_idx in range(num_batches):
    print(f"Processing batch {batch_idx + 1} of {num_batches}")
    start = batch_idx * save_batch
    end = min(start + save_batch, total_rows)

    batch_rows = []
    for i, row in df.iloc[start:end].iterrows():
        output = zeroshot_classifier(
            row["Description"],
            classes,
            hypothesis_template=hypothesis_template,
            multi_label=True
        )
        scores_dict = dict(zip(output["labels"], output["scores"]))
        row_result = {label: scores_dict.get(label, 0.0) for label in classes}
        row_result["ID expérience"] = row["ID expérience"]
        row_result["Description"] = row["Description"]
        batch_rows.append(row_result)

        # Convert all results to DataFrame
    df_batch = pd.DataFrame(batch_rows)[["ID expérience", "Description"] + classes]

    # Save to Excel (append or create)
    if batch_idx == 0 or not os.path.exists(excel_file):
        # First batch: write with header
        df_batch.to_excel(excel_file, index=False, engine="openpyxl")
    else:
        # Append without writing header
        with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            # Get existing rows in sheet
            book = writer.book
            sheet = book.active
            is_empty = (sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None)
            startrow = 0 if is_empty else sheet.max_row

            # Write below existing rows
            df_batch.to_excel(writer, index=False, header=False, startrow=startrow)

            # if device.type == "mps":
            #     torch.mps.empty_cache()
            # gc.collect()

    print(f"Saved batch {batch_idx + 1} to {excel_file}")

# Enregistrer le temps de fin
end_time = time.time()

# Calculer le temps d'exécution
execution_time = end_time - start_time
print(f"Temps d'exécution: {execution_time} secondes")

times2 = pd.DataFrame([{
    "Modèle": m,
    "Taille de lot": batch_size,
    "Execution (s)": round(execution_time, 2),
}])

times=pd.concat([times, times2], ignore_index=True)

push_to_github("Roberta")

####################### Troisieme modele ############################

print('modele camembert')
m='camembert'
excel_file = os.path.join(output_folder, f"{m}.xlsx")

zeroshot_classifier = pipeline("zero-shot-classification", model="mtheo/camembert-base-xnli",
                               device=device, batch_size=batch_size, multilabel=True)
hypothesis_template = "The emotions of this text are {}"



# Enregistrer le temps de début
start_time = time.time()
print(start_time)



for batch_idx in range(num_batches):
    print(f"Processing batch {batch_idx + 1} of {num_batches}")
    start = batch_idx * save_batch
    end = min(start + save_batch, total_rows)

    batch_rows = []
    for i, row in df.iloc[start:end].iterrows():
        output = zeroshot_classifier(
            row["Description"],
            classes,
            hypothesis_template=hypothesis_template,
            multi_label=True
        )
        scores_dict = dict(zip(output["labels"], output["scores"]))
        row_result = {label: scores_dict.get(label, 0.0) for label in classes}
        row_result["ID expérience"] = row["ID expérience"]
        row_result["Description"] = row["Description"]
        batch_rows.append(row_result)

    # Convert all results to DataFrame
    df_batch = pd.DataFrame(batch_rows)[["ID expérience", "Description"] + classes]  # FIXED


    # Save to Excel (append or create)
    if batch_idx == 0 or not os.path.exists(excel_file):
        # First batch: write with header
        df_batch.to_excel(excel_file, index=False, engine="openpyxl")
    else:
        # Append without writing header
        with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            # Get existing rows in sheet
            book = writer.book
            sheet = book.active
            startrow = sheet.max_row

            # Write below existing rows
            df_batch.to_excel(writer, index=False, header=False, startrow=startrow)

            # if device.type == "mps":
            #     torch.mps.empty_cache()
            # gc.collect()

    print(f"Saved batch {batch_idx + 1} to {excel_file}")

# Enregistrer le temps de fin
end_time = time.time()

# Calculer le temps d'exécution
execution_time = end_time - start_time
print(f"Temps d'exécution: {execution_time} secondes")

times2 = pd.DataFrame([{
    "Modèle": m,
    "Taille de lot": batch_size,
    "Execution (s)": round(execution_time, 2),
}])
times=pd.concat([times, times2], ignore_index=True)

times.to_csv(os.path.join(output_folder, "times.csv"), index=False)


push_to_github("Camembert")
